import locale
from typing import Any

import openpyxl

from django.contrib import admin
from django.db.models.query import QuerySet
from django.http import HttpRequest, HttpResponse

from base import models, forms


try:
    locale.setlocale(locale.LC_TIME, 'ru_RU.UTF-8')
except locale.Error:
    print("Не удалось установить локаль, используем локаль по умолчанию.")


class CustomUserAdmin(admin.ModelAdmin):
    list_display = ["username"]
    exclude = ["id", "user_permissions"]
    search_fields = ["username", "id"]
    readonly_fields = ["last_login", "is_superuser"]

    add_form = forms.CustomUserCreationForm

    def get_fieldsets(self, request, obj=None):
        if obj is None:
            return [(None, {"fields": ("username", "password")})]
        return super().get_fieldsets(request, obj)

    def add_view(self, request, form_url="", extra_context=None):
        self.fieldsets = [(None, {"fields": ("username", "password")})]
        return super().add_view(request, form_url, extra_context)
    
    def change_view(self, request, object_id, form_url='', extra_context=None):
        self.fieldsets = None
        return super().change_view(request, object_id, form_url, extra_context)

    def save_model(self, request, obj, form, change):
        if not change:
            obj.set_password(obj.password)
        super().save_model(request, obj, form, change)
        
        
admin.site.register(models.User, CustomUserAdmin)


# Inlines
class StorageClientM2MInline(admin.TabularInline):
    model = models.StorageClientM2M
    extra = 1


class ItemImageInline(admin.TabularInline):
    model = models.ItemImage
    max_num = 5
    fields = ["image_tag", "image"]
    readonly_fields = ["image_tag"]
    
    def get_min_num(self, request, obj=None, **kwargs):
        if self.parent_model == models.ItemStock:
            return 0
        return 1

    def get_extra(self, request, obj=None, **kwargs):
        if self.parent_model == models.ItemStock:
            return 5
        return 4

    def has_change_permission(self, request, obj=None):
        if obj and hasattr(obj, 'is_approved') and obj.is_approved:
            return False
        return super().has_change_permission(request, obj)

    def has_add_permission(self, request, obj=None):
        if obj and hasattr(obj, 'is_approved') and obj.is_approved:
            return False
        return super().has_add_permission(request, obj)

    def has_delete_permission(self, request, obj=None):
        if obj and hasattr(obj, 'is_approved') and obj.is_approved:
            return False
        return super().has_delete_permission(request, obj)


class ItemBookingItemM2MInline(admin.TabularInline):
    model = models.ItemBookingItemM2M
    min_num = 1
    extra = 0
    validate_min = True
    
    class Media:
        js = (
            "//ajax.googleapis.com/ajax/libs/jquery/3.4.1/jquery.min.js",
            "admin/js/check_item_booking.js",
        )


class RecoveryImageInline(admin.TabularInline):
    model = models.RecoveryImage
    extra = 4
    min_num = 1
    max_num = 5
    validate_min = True
    fields = ["image_tag", "image"]
    readonly_fields = ["image_tag"]


class RefundImageInline(admin.TabularInline):
    model = models.ItemRefundImage
    extra = 4
    min_num = 1
    max_num = 5
    validate_min = True
    fields = ["image_tag", "image"]
    readonly_fields = ["image_tag"]


class ItemRefundItemM2MInline(admin.TabularInline):
    model = models.ItemRefundItemM2M
    min_num = 1
    extra = 0
    validate_min = True


class ItemConsumptionImageInline(admin.TabularInline):
    model = models.ItemConsumptionImage
    min_num = 1
    extra = 4
    max_num = 5
    validate_min = True
    fields = ["image_tag", "image"]
    readonly_fields = ["image_tag"]


# Models
@admin.register(models.Client)
class ClientAdmin(admin.ModelAdmin):
    list_display = ["name"]
    exclude = ["id"]
    search_fields = ["name"]
    inlines = [StorageClientM2MInline]


@admin.register(models.Storage)
class StorageAdmin(admin.ModelAdmin):
    list_display = ["name", "free_area", "area"]
    exclude = ["id"]
    search_fields = ["name", "clients"]
    
    def get_queryset(self, request: HttpRequest) -> QuerySet:
        if request.user.groups.filter(name="Кладовщик").exists():
            queryset = request.user.storages.all()
            return queryset
        return super().get_queryset(request)


@admin.register(models.Project)
class ProjectAdmin(admin.ModelAdmin):
    list_display = ["name", "client"]
    exclude = ["id"]
    search_fields = ["name", "client"]


@admin.register(models.ItemCategory)
class ItemCategoryAdmin(admin.ModelAdmin):
    list_display = ["name"]
    exclude = ["id"]
    search_fields = ["name"]


@admin.register(models.ItemStatus)
class ItemStatusAdmin(admin.ModelAdmin):
    list_display = ["text"]
    exclude = ["id"]


@admin.register(models.Item)
class ItemAdmin(admin.ModelAdmin):
    list_display = ["article", "name", "category", "count", "is_booked"]
    exclude = ["id"]
    search_fields = ["article", "name"]
    readonly_fields = ["article", "is_booked", "booking_projects", "booking_quantities", "booking_periods"]
    inlines = [ItemImageInline]
    actions = ["export_as_xlsx"]
    
    def export_as_xlsx(self, request, queryset):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Товары"

        headers = ['Артикул', 'Название', 'Категория', 'Количество на складе', 'Забронирован?', 'Периоды броней']
        sheet.append(headers)

        for obj in queryset:
            bookings = models.ItemBooking.objects.filter(items=obj)
            periods = ", ".join(
                f"{booking.start_date.strftime('%d.%m.%Y')}-{booking.end_date.strftime('%d.%m.%Y')}" 
                for booking in bookings
            )
            sheet.append([
                obj.article,
                obj.name,
                obj.category.name if obj.category else None,
                obj.count,
                'Да' if obj.is_booked else 'Нет',
                periods or "Нет броннирований",
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=товары.xlsx'

        workbook.save(response)
        return response

    export_as_xlsx.short_description = "Выгрузить .XLSX"

    @admin.display(description="Проекты")
    def booking_projects(self, obj):
        projects = obj.bookings.values_list('project__name', flat=True).distinct()
        return ", ".join(projects) if projects else "—"

    @admin.display(description="Количество")
    def booking_quantities(self, obj):
        booking_items = models.ItemBookingItemM2M.objects.filter(item=obj)
        quantities = [bi.item_count for bi in booking_items]
        return sum(quantities) if quantities else "—"

    @admin.display(description="Периоды брони")
    def booking_periods(self, obj):
        periods = obj.bookings.values_list('start_date', 'end_date')
        return " | ".join([
            f"{start.strftime('%d %B %Y')} - {end.strftime('%d %B %Y')}" for start, end in periods
        ]) if periods else "—"


@admin.register(models.ItemStock)
class AdminItemStock(admin.ModelAdmin):
    list_display = ["article_display", "client_display", "storage_display", "count", "is_archived"]
    list_filter = ('request_type', "is_archived")
    search_fields = ('new_item_name', 'existing_item__name')
    inlines = [ItemImageInline]
    actions = ["export_as_xlsx"]
    
    def export_as_xlsx(self, request, queryset):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Заявки на приход"

        headers = ["Товар", "Проект", "Дата прихода"]
        sheet.append(headers)

        for obj in queryset:
            sheet.append([
                f"{obj.new_item_name} ({obj.new_item_count}шт.) {obj.new_item_storage.name}"
                if obj.new_item_name
                else f"{obj.existing_item.name} ({obj.count}шт.) {obj.existing_item.storage.name}",
                obj.existing_item.project.name if obj.existing_item and obj.existing_item.project else obj.new_item_project,
                obj.date.strftime('%d.%m.%Y'),
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=товары.xlsx'

        workbook.save(response)
        return response

    export_as_xlsx.short_description = "Выгрузить .XLSX"
    
    @admin.display(description="Клиент")
    def client_display(self, obj):
        if obj.existing_item:
            if obj.existing_item.client:
                return obj.existing_item.client
            else:
                if obj.existing_item.project:
                    return obj.existing_item.project.client
        else:
            return obj.new_item_client
    
    @admin.display(description="Склад")
    def storage_display(self, obj):
        if obj.existing_item:
            return obj.existing_item.storage
        else:
            return obj.new_item_storage
    
    @admin.display(description="Товар")
    def article_display(self, obj):
        if obj.existing_item:
            return obj.existing_item
        else:
            return obj.new_item_name
    
    def get_readonly_fields(self, request: HttpRequest, obj: Any | None = ...) -> list[str] | tuple[Any, ...]:
        if obj and obj.is_approved:
            return [field.name for field in obj._meta.fields]
        
        if request.user.groups.filter(name="Кладовщик").exists():
            fields = [
                "request_type", "existing_item", #"count",
                "new_item_name", "new_item_description",
                "new_item_weight", "new_item_height",
                "new_item_width", "new_item_length",
                "new_item_project", "new_item_client",
                "new_item_storage", "new_item_category",
                "new_item_status", "new_item_arrival_date",
                "new_item_expiration_date", "planning_date",
                "is_archived",
            ]
            return fields
        else:
            return ["is_approved", "date", "is_archived",]
    
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.order_by("is_archived")
    
    class Media:
        js = (
            "//ajax.googleapis.com/ajax/libs/jquery/3.4.1/jquery.min.js",
            "admin/js/admin_itemstock.js",
        )


@admin.register(models.ItemBooking)
class AdminItemBooking(admin.ModelAdmin):
    list_display = ('project', 'city', 'is_approved', 'booking_items', 'booking_quantities', 'booking_periods', "is_archived")
    form = forms.BookingAdminForm
    exclude = ["id"]
    search_fields = ["project__name", "items__name", "start_date__month"] # TODO: add month
    inlines = [ItemBookingItemM2MInline]
    list_filter = ["is_archived"]
    actions = ["export_as_xlsx"]
    
    def export_as_xlsx(self, request, queryset):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Заявки на бронь"

        headers = ["Товары", "Проект", "Город", "Начальная дата", "Конечная дата"]
        sheet.append(headers)

        for obj in queryset:
            items = obj.items.all()
            items_data = ", ".join(
                f"{item.article} | {item.name} ({item.count}шт.) [{item.storage.name if item.storage else 'Склад не указан'}]"
                for item in items
            )
            sheet.append([
                items_data,
                obj.project.name if obj.project else None,
                obj.city,
                obj.start_date.strftime('%d.%m.%Y'),
                obj.end_date.strftime('%d.%m.%Y'),
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=товары.xlsx'

        workbook.save(response)
        return response

    export_as_xlsx.short_description = "Выгрузить .XLSX"
    
    @admin.display(description="Товары")
    def booking_items(self, obj):
        items = obj.items.values_list('name', flat=True)
        return ", ".join(items) if items else "—"

    @admin.display(description="Количество")
    def booking_quantities(self, obj):
        booking_items = models.ItemBookingItemM2M.objects.filter(booking=obj)
        quantities = [f"{bi.item.name}: {bi.item_count}" for bi in booking_items]
        return ", ".join(quantities) if quantities else "—"

    @admin.display(description="Периоды брони")
    def booking_periods(self, obj):
        start_date = obj.start_date.strftime('%d %B %Y') if obj.start_date else "—"
        end_date = obj.end_date.strftime('%d %B %Y') if obj.end_date else "—"
        return f"{start_date} - {end_date}"
    
    def get_readonly_fields(self, request: HttpRequest, obj: Any | None = ...) -> list[str] | tuple[Any, ...]:
        if request.user.groups.filter(name="Кладовщик").exists():
            fields = [
                "items", "project", "date",
                "city", "description", "start_date",
                "end_date", "is_archived",
            ]
            return fields
        else:
            return ["is_approved", "is_archived",]
    
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.order_by("is_archived")
    
    
@admin.register(models.ItemRecovery)
class AdminItemRecovery(admin.ModelAdmin):
    list_display = ["item", "count", "item__storage", "planning_date", "is_ceo_approved", "is_approved", "is_archived"]
    exclude = ["id"]
    search_fields = ["item__article", "item__name"]
    inlines = [RecoveryImageInline]
    list_filter = ["is_archived"]
    actions = ["export_as_xlsx"]
    
    def export_as_xlsx(self, request, queryset):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Заявки на утилизацию"

        headers = ["Товар", "Количество", "Планируемая дата", "Подтверждение утилизации кладовщиком"]
        sheet.append(headers)

        for obj in queryset:
            sheet.append([
                f"{obj.item.article} | {obj.item.name}",
                obj.item.count,
                obj.planning_date.strftime("%d.%m.%Y"),
                "Да" if obj.is_approved else "Нет",
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=товары.xlsx'

        workbook.save(response)
        return response

    export_as_xlsx.short_description = "Выгрузить .XLSX"
    
    def get_readonly_fields(self, request: HttpRequest, obj: Any | None = ...) -> list[str] | tuple[Any, ...]:
        if request.user.groups.filter(name="Кладовщик").exists():
            fields = [
                "item", "reason", "planning_date",
                "description", "status", "is_ceo_approved",
                "is_archived",
            ]
            return fields
        elif request.user.groups.filter(name="Руководитель").exists():
            return ["is_approved", "date", "is_archived",]
        else:
            return ["is_approved", "is_ceo_approved", "date", "is_archived",]
    
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.order_by("is_archived")

    
@admin.register(models.ItemRefund)
class AdminItemRefund(admin.ModelAdmin):
    exclude = ["id"]
    search_fields = ["items__article", "items__name"]
    inlines = [ItemRefundItemM2MInline, RefundImageInline]
    list_display = ["project__name", "project__client", "city", "date", "storages_display", "is_archived"]
    list_filter = ["is_archived"]
    actions = ["export_as_xlsx"]
    
    def export_as_xlsx(self, request, queryset):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Заявки на возврат"

        headers = ["Товары", "Проект", "Город (откуда едет)", "Дата прихода"]
        sheet.append(headers)

        for obj in queryset:
            items = obj.items.all()
            items_data = ", ".join(
                f"{item.article} | {item.name} ({item.count}шт.) [{item.storage.name if item.storage else 'Склад не указан'}]"
                for item in items
            )
            sheet.append([
                items_data,
                obj.project.name,
                obj.city,
                obj.date.strftime('%d.%m.%Y'),
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=товары.xlsx'

        workbook.save(response)
        return response

    export_as_xlsx.short_description = "Выгрузить .XLSX"
    
    @admin.display(description="Склады")
    def storages_display(self, obj):
        return " | ".join(set([item.storage.name for item in obj.items.all()]))
    
    def get_readonly_fields(self, request: HttpRequest, obj: Any | None = ...) -> list[str] | tuple[Any, ...]:
        if request.user.groups.filter(name="Кладовщик").exists():
            fields = [
                "items", "project", #"description",
                "is_archived",
                # "status",
            ]
            return fields
        else:
            return ["is_approved", "is_archived",]
    
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.order_by("is_archived")
    

@admin.register(models.ItemConsumption)
class AdminItemConsumption(admin.ModelAdmin):
    list_display = ["booking__project__name", "booking__project__client", "city", "date_display", "storage_display", "is_archived"]
    exclude = ["id"]
    search_fields = ["booking__items__article", "booking__items__name", "date__month"]
    inlines = [ItemConsumptionImageInline]
    list_filter = ["is_archived"]
    actions = ["export_as_xlsx"]
    
    def export_as_xlsx(self, request, queryset):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Заявки на расход"

        headers = ["Товары", "Проект", "Город (куда едет)", "Дата отправки"]
        sheet.append(headers)

        for obj in queryset:
            items = obj.booking.items.all()
            items_data = ", ".join(
                f"{item.article} | {item.name} ({item.count}шт.) [{item.storage.name if item.storage else 'Склад не указан'}]"
                for item in items
            )
            sheet.append([
                items_data,
                obj.booking.project.name if obj.booking.project else None,
                obj.city,
                obj.date.strftime('%d.%m.%Y'),
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=товары.xlsx'

        workbook.save(response)
        return response

    export_as_xlsx.short_description = "Выгрузить .XLSX"
    
    @admin.display(description="Дата отправки")
    def date_display(self, obj):
        return obj.date.strftime('%d %B %Y') if obj.date else ""
    
    @admin.display(description="Склады")
    def storage_display(self, obj):
        storages = set([item.storage.name for item in obj.booking.items.all()])
        return ", ".join(storages)
    
    def get_readonly_fields(self, request: HttpRequest, obj: Any | None = ...) -> list[str] | tuple[Any, ...]:
        if request.user.groups.filter(name="Кладовщик").exists():
            return [
                "booking", "date_created", "is_archived", #"description",
            ]
        else:
            return ["date_created", "is_approved", "is_archived",]
    
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.order_by("is_archived")
    
    